import pywhatkit as pkit
from openpyxl import Workbook, load_workbook

database = load_workbook("database.xlsx")
data = database.active


for i in range(2,len(data.cell(1,1).value +"1")):
    Name = data["A"+str(i)].value
    Number = data["B"+str(i)].value
    Tıme = data["C"+str(i)].value
    Message = data["D"+str(i)].value
    information = (Name,Number,Tıme,Message)
    print(information)